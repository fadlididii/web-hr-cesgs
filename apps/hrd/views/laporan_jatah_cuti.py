from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_POST
from apps.hrd.models import Karyawan, JatahCuti, DetailJatahCuti, CutiBersama
from django.db.models import Q
import calendar
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
from datetime import datetime, timedelta
from django.core.paginator import Paginator
import json

@login_required
def laporan_jatah_cuti_view(request):
    """View untuk menampilkan laporan jatah cuti per karyawan per bulan."""
    if request.user.role != 'HRD':
        messages.error(request, "Anda tidak memiliki akses ke halaman ini.")
        return redirect('karyawan_dashboard')
    
    # Filter berdasarkan tahun
    tahun = request.GET.get('tahun', datetime.now().year)
    try:
        tahun = int(tahun)
    except (ValueError, TypeError):
        tahun = datetime.now().year
    
    # Filter berdasarkan nama karyawan (opsional)
    nama = request.GET.get('nama', '')
    
    # Ambil semua karyawan tetap dan HRD yang aktif
    karyawan_list = Karyawan.objects.filter(
        Q(user__role='HRD') | Q(user__role='Karyawan Tetap'),
        status_keaktifan='Aktif'
    ).order_by('nama')
    
    if nama:
        karyawan_list = karyawan_list.filter(nama__icontains=nama)
    
    # Data untuk laporan
    laporan_data = []
    
    # Cek cuti yang expired
    cuti_expired = []
    current_date = datetime.now().date()
    
    for karyawan in karyawan_list:
        # Ambil jatah cuti untuk tahun yang dipilih
        jatah_cuti = JatahCuti.objects.filter(karyawan=karyawan, tahun=tahun).first()
        
        if not jatah_cuti:
            # Jika belum ada jatah cuti, buat data kosong
            bulan_data = [{'dipakai': False, 'keterangan': '', 'expired': False} for _ in range(1, 13)]
            total_cuti = 0
            sisa_cuti = 0
        else:
            # Ambil detail jatah cuti per bulan
            detail_cuti = DetailJatahCuti.objects.filter(
                jatah_cuti=jatah_cuti,
                tahun=tahun
            ).order_by('bulan')
            
            # Buat dictionary untuk memetakan bulan ke data cuti
            bulan_data = {}
            for detail in detail_cuti:
                # Cek apakah cuti sudah expired (lebih dari 1 tahun)
                expired = False
                if not detail.dipakai:
                    # Hitung tanggal expired (1 tahun dari bulan tersebut)
                    # Gunakan calendar.monthrange untuk tahun+1 dan bulan yang sesuai
                    # untuk mendapatkan jumlah hari yang benar di bulan tersebut pada tahun berikutnya
                    last_day_of_month = calendar.monthrange(tahun + 1, detail.bulan)[1]
                    expired_date = datetime(tahun + 1, detail.bulan, last_day_of_month).date()
                    
                    if current_date > expired_date:
                        expired = True
                        # Tambahkan ke daftar cuti expired untuk notifikasi
                        if expired:
                            cuti_expired.append({
                                'karyawan': karyawan.nama,
                                'bulan': calendar.month_name[detail.bulan],
                                'tahun': tahun
                            })
                
                bulan_data[detail.bulan] = {
                    'dipakai': detail.dipakai,
                    'keterangan': detail.keterangan,
                    'expired': expired
                }
            
            # Pastikan semua bulan ada datanya
            bulan_data = [bulan_data.get(bulan, {'dipakai': False, 'keterangan': '', 'expired': False}) for bulan in range(1, 13)]
            total_cuti = jatah_cuti.total_cuti
            sisa_cuti = jatah_cuti.sisa_cuti
        
        # Tambahkan data karyawan ke laporan
        laporan_data.append({
            'karyawan': karyawan,
            'bulan_data': bulan_data,
            'total_cuti': total_cuti,
            'sisa_cuti': sisa_cuti
        })
    
    # Nama-nama bulan untuk header tabel
    nama_bulan = [calendar.month_name[i] for i in range(1, 13)]
    
    return render(request, 'hrd/laporan_jatah_cuti.html', {
        'laporan_data': laporan_data,
        'nama_bulan': nama_bulan,
        'tahun': tahun,
        'nama': nama,
        'cuti_expired': cuti_expired,
        'tahun_options': range(datetime.now().year - 2, datetime.now().year + 3)  # Opsi tahun: 2 tahun ke belakang, tahun ini, 2 tahun ke depan
    })

@login_required
def export_laporan_jatah_cuti_excel(request):
    """View untuk mengekspor laporan jatah cuti ke Excel."""
    if request.user.role != 'HRD':
        return HttpResponse("Forbidden", status=403)
    
    # Filter berdasarkan tahun
    tahun = request.GET.get('tahun', datetime.now().year)
    try:
        tahun = int(tahun)
    except (ValueError, TypeError):
        tahun = datetime.now().year
    
    # Filter berdasarkan nama karyawan (opsional)
    nama = request.GET.get('nama', '')
    
    # Ambil semua karyawan tetap dan HRD yang aktif
    karyawan_list = Karyawan.objects.filter(
        Q(user__role='HRD') | Q(user__role='Karyawan Tetap'),
        status_keaktifan='Aktif'
    ).order_by('nama')
    
    if nama:
        karyawan_list = karyawan_list.filter(nama__icontains=nama)
    
    # Buat workbook Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Laporan Jatah Cuti {tahun}"
    
    # Styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    expired_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    used_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Header
    headers = ["No", "Nama Lengkap", "Januari", "Februari", "Maret", "April", "Mei", "Juni", 
              "Juli", "Agustus", "September", "Oktober", "November", "Desember", "Saldo Cuti"]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    row_num = 2
    current_date = datetime.now().date()
    
    for idx, karyawan in enumerate(karyawan_list, 1):
        # Ambil jatah cuti untuk tahun yang dipilih
        jatah_cuti = JatahCuti.objects.filter(karyawan=karyawan, tahun=tahun).first()
        
        if not jatah_cuti:
            # Jika belum ada jatah cuti, buat data kosong
            bulan_data = ['' for _ in range(1, 13)]
            sisa_cuti = 0
        else:
            # Ambil detail jatah cuti per bulan
            detail_cuti = DetailJatahCuti.objects.filter(
                jatah_cuti=jatah_cuti,
                tahun=tahun
            ).order_by('bulan')
            
            # Buat dictionary untuk memetakan bulan ke data cuti
            bulan_data = {}
            for detail in detail_cuti:
                # Cek apakah cuti sudah expired (lebih dari 1 tahun)
                expired = False
                if not detail.dipakai:
                    # Hitung tanggal expired (1 tahun dari bulan tersebut)
                    # Gunakan calendar.monthrange untuk tahun+1 dan bulan yang sesuai
                    # untuk mendapatkan jumlah hari yang benar di bulan tersebut pada tahun berikutnya
                    last_day_of_month = calendar.monthrange(tahun + 1, detail.bulan)[1]
                    expired_date = datetime(tahun + 1, detail.bulan, last_day_of_month).date()
                    
                    if current_date > expired_date:
                        expired = True
                        # Tambahkan ke daftar cuti expired untuk notifikasi
                        if expired:
                            cuti_expired.append({
                                'karyawan': karyawan.nama,
                                'bulan': calendar.month_name[detail.bulan],
                                'tahun': tahun
                            })
                
                if detail.dipakai:
                    tanggal = detail.keterangan.split(': ')[-1] if ': ' in detail.keterangan else ''
                    bulan_data[detail.bulan] = {'text': tanggal, 'dipakai': True, 'expired': False}
                else:
                    bulan_data[detail.bulan] = {'text': '', 'dipakai': False, 'expired': expired}
            
            # Pastikan semua bulan ada datanya
            bulan_data = [bulan_data.get(bulan, {'text': '', 'dipakai': False, 'expired': False}) for bulan in range(1, 13)]
            sisa_cuti = jatah_cuti.sisa_cuti
        
        # Tambahkan data ke Excel
        row = [idx, karyawan.nama]
        
        # Tambahkan data bulan
        for bulan_info in bulan_data:
            row.append(bulan_info['text'])
        
        # Tambahkan sisa cuti
        row.append(sisa_cuti)
        
        # Tulis ke Excel
        for col_num, value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            
            # Styling untuk bulan
            if 2 < col_num < len(headers):
                cell.alignment = Alignment(horizontal='center')
                
                # Jika bulan dipakai, beri warna hijau
                if bulan_data[col_num-3]['dipakai']:
                    cell.fill = used_fill
                
                # Jika bulan expired, beri warna merah
                if bulan_data[col_num-3]['expired']:
                    cell.fill = expired_fill
        
        row_num += 1
    
    # Atur lebar kolom
    for col_num in range(1, len(headers) + 1):
        if col_num == 2:  # Kolom nama
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 30
        else:
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 15
    
    # Freeze panes
    ws.freeze_panes = 'C2'
    
    # Buat response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=laporan_jatah_cuti_{tahun}.xlsx'
    wb.save(response)
    
    return response


@login_required
@require_POST
def update_jatah_cuti_ajax(request):
    """View untuk memperbarui detail jatah cuti via AJAX."""
    if request.user.role != 'HRD':
        return JsonResponse({'status': 'error', 'message': 'Forbidden'}, status=403)
    
    try:
        data = json.loads(request.body)
        karyawan_id = data.get('karyawan_id')
        tahun = data.get('tahun')
        bulan = data.get('bulan')
        dipakai = data.get('dipakai', False)
        keterangan = data.get('keterangan', '')
        
        # Validasi input
        if not all([karyawan_id, tahun, bulan]):
            return JsonResponse({'status': 'error', 'message': 'Data tidak lengkap'}, status=400)
        
        # Ambil karyawan
        karyawan = get_object_or_404(Karyawan, id=karyawan_id)
        
        # Ambil atau buat jatah cuti
        jatah_cuti, created = JatahCuti.objects.get_or_create(
            karyawan=karyawan,
            tahun=int(tahun),
            defaults={
                'total_cuti': 12,
                'sisa_cuti': 12
            }
        )
        
        # Ambil atau buat detail jatah cuti
        detail, created = DetailJatahCuti.objects.get_or_create(
            jatah_cuti=jatah_cuti,
            tahun=int(tahun),
            bulan=int(bulan),
            defaults={
                'dipakai': False,
                'jumlah_hari': 0,
                'keterangan': ''
            }
        )
        
        # Simpan status sebelumnya untuk menghitung perubahan sisa cuti
        previous_dipakai = detail.dipakai
        
        # Update detail
        detail.dipakai = dipakai
        detail.jumlah_hari = 1 if dipakai else 0
        detail.keterangan = keterangan
        detail.save()
        
        # Update sisa cuti
        if dipakai and not previous_dipakai:
            # Jika sebelumnya tidak dipakai dan sekarang dipakai, kurangi sisa cuti
            jatah_cuti.sisa_cuti = max(0, jatah_cuti.sisa_cuti - 1)
        elif not dipakai and previous_dipakai:
            # Jika sebelumnya dipakai dan sekarang tidak dipakai, tambah sisa cuti
            jatah_cuti.sisa_cuti += 1
        
        jatah_cuti.save()
        
        # Cek apakah cuti sudah expired
        expired = False
        if not dipakai:
            current_date = datetime.now().date()
            last_day_of_month = calendar.monthrange(int(tahun), int(bulan))[1]
            expired_date = datetime(int(tahun) + 1, int(bulan), last_day_of_month).date()
            
            if current_date > expired_date:
                expired = True
        
        return JsonResponse({
            'status': 'success',
            'data': {
                'dipakai': detail.dipakai,
                'keterangan': detail.keterangan,
                'sisa_cuti': jatah_cuti.sisa_cuti,
                'expired': expired
            }
        })
    
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@login_required
def get_detail_jatah_cuti_ajax(request):
    """View untuk mendapatkan detail jatah cuti via AJAX."""
    if request.user.role != 'HRD':
        return JsonResponse({'status': 'error', 'message': 'Forbidden'}, status=403)
    
    try:
        karyawan_id = request.GET.get('karyawan_id')
        tahun = request.GET.get('tahun')
        bulan = request.GET.get('bulan')
        
        # Validasi input
        if not all([karyawan_id, tahun, bulan]):
            return JsonResponse({'status': 'error', 'message': 'Data tidak lengkap'}, status=400)
        
        # Ambil karyawan
        karyawan = get_object_or_404(Karyawan, id=karyawan_id)
        
        # Ambil jatah cuti
        jatah_cuti = JatahCuti.objects.filter(
            karyawan=karyawan,
            tahun=int(tahun)
        ).first()
        
        if not jatah_cuti:
            return JsonResponse({
                'status': 'success',
                'data': {
                    'dipakai': False,
                    'keterangan': '',
                    'sisa_cuti': 0,
                    'expired': False
                }
            })
        
        # Ambil detail jatah cuti
        detail = DetailJatahCuti.objects.filter(
            jatah_cuti=jatah_cuti,
            tahun=int(tahun),
            bulan=int(bulan)
        ).first()
        
        # Cek apakah cuti sudah expired
        expired = False
        if detail and not detail.dipakai:
            current_date = datetime.now().date()
            last_day_of_month = calendar.monthrange(int(tahun), int(bulan))[1]
            expired_date = datetime(int(tahun) + 1, int(bulan), last_day_of_month).date()
            
            if current_date > expired_date:
                expired = True
        
        if not detail:
            return JsonResponse({
                'status': 'success',
                'data': {
                    'dipakai': False,
                    'keterangan': '',
                    'sisa_cuti': jatah_cuti.sisa_cuti,
                    'expired': expired
                }
            })
        
        return JsonResponse({
            'status': 'success',
            'data': {
                'dipakai': detail.dipakai,
                'keterangan': detail.keterangan,
                'sisa_cuti': jatah_cuti.sisa_cuti,
                'expired': expired
            }
        })
    
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)