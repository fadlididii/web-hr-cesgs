from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from apps.hrd.models import JatahCuti, DetailJatahCuti
from django.http import HttpResponse, JsonResponse
from datetime import datetime
from django.views.decorators.http import require_POST
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Alignment

@login_required
def lihat_jatah_cuti_view(request):
    if request.user.role != 'HRD':
        messages.error(request, "Anda tidak memiliki akses ke halaman ini.")
        return redirect('karyawan_dashboard')

    tahun = int(request.GET.get('tahun') or datetime.now().year)
    nama = request.GET.get('nama')

    queryset = JatahCuti.objects.select_related('karyawan').prefetch_related('detail')
    if tahun:
        queryset = queryset.filter(tahun=tahun)
    if nama:
        queryset = queryset.filter(karyawan__nama__icontains=nama)

    data = []
    for jatah in queryset.order_by('karyawan__nama'):
        # Perhitungan sisa cuti tidak menghitung bulan yang tidak aktif
        sisa_cuti = jatah.detail.filter(dipakai=False).exclude(keterangan__icontains='tidak aktif').count()
        if jatah.sisa_cuti != sisa_cuti:
            jatah.sisa_cuti = sisa_cuti
            jatah.save()

        bulanan = {i: '-' for i in range(1, 13)}
        for detail in jatah.detail.all():
            if 'Belum aktif' in detail.keterangan:
                bulanan[detail.bulan] = 'belum_aktif'
            elif 'Hangus Carry' in detail.keterangan:
                bulanan[detail.bulan] = 'hangus_carry'
            elif 'Carry Forward' in detail.keterangan:
                bulanan[detail.bulan] = 'carry'
            elif not detail.dipakai:
                bulanan[detail.bulan] = '-'
            elif 'Cuti Tahunan' in detail.keterangan:
                bulanan[detail.bulan] = 'cuti'
            elif 'Cuti Bersama' in detail.keterangan:
                bulanan[detail.bulan] = 'cuti_bersama'
            elif 'Hangus' in detail.keterangan:
                bulanan[detail.bulan] = 'hangus'
            else:
                bulanan[detail.bulan] = 'cuti'

        data.append({
            'id': jatah.id,
            'nama': jatah.karyawan.nama,
            'tahun': jatah.tahun,
            'total_cuti': jatah.total_cuti,
            'sisa_cuti': jatah.sisa_cuti,
            'bulanan': bulanan,
        })

    bulan_list = ['Jan', 'Feb', 'Mar', 'Apr', 'Mei', 'Jun', 'Jul', 'Agu', 'Sep', 'Okt', 'Nov', 'Des']

    return render(request, 'hrd/jatah_cuti.html', {
        'data': data,
        'tahun': tahun,
        'bulan_list': bulan_list,
    })

@login_required
def edit_jatah_cuti_view(request):
    if request.user.role != 'HRD':
        messages.error(request, "Akses ditolak.")
        return redirect('karyawan_dashboard')

    if request.method == 'POST':
        cuti_id = request.POST.get('id')
        total = int(request.POST.get('total_cuti'))
        sisa = int(request.POST.get('sisa_cuti'))

        jatah = JatahCuti.objects.get(id=cuti_id)
        jatah.total_cuti = total
        jatah.sisa_cuti = sisa
        jatah.save()

        messages.success(request, "Jatah cuti berhasil diperbarui.")
    return redirect('lihat_jatah_cuti')

@login_required
def export_jatah_cuti_excel(request):
    tahun = int(request.GET.get('tahun', datetime.now().year))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Cuti {tahun}"

    ws['B2'] = f"Tahun {tahun}"
    ws['B3'] = "Nama Karyawan"
    for bulan in range(1, 13):
        ws.cell(row=3, column=bulan+2).value = bulan

    queryset = JatahCuti.objects.filter(tahun=tahun).select_related('karyawan').prefetch_related('detail')
    for row_idx, jatah in enumerate(queryset.order_by('karyawan__nama'), start=4):
        ws.cell(row=row_idx, column=2).value = jatah.karyawan.nama
        for detail in jatah.detail.all():
            col = detail.bulan + 2
            if detail.dipakai:
                if 'Cuti Tahunan' in detail.keterangan:
                    fill_color = "FFFF00"
                elif 'Cuti Bersama' in detail.keterangan:
                    fill_color = "FF6666"
                elif 'Hangus' in detail.keterangan:
                    fill_color = "CCCCCC"
                else:
                    fill_color = "AAAAAA"
                ws.cell(row=row_idx, column=col).value = 1
                ws.cell(row=row_idx, column=col).fill = PatternFill("solid", fgColor=fill_color)

    for col in ws.columns:
        for cell in col:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Jatah_Cuti_{tahun}.xlsx"'
    wb.save(response)
    return response

@login_required
def grid_jatah_cuti_view(request):
    if request.user.role != 'HRD':
        messages.error(request, "Anda tidak memiliki akses.")
        return redirect('karyawan_dashboard')

    tahun = request.GET.get('tahun', datetime.now().year)
    data = JatahCuti.objects.filter(tahun=tahun).select_related('karyawan').prefetch_related('detail').order_by('karyawan__nama')

    return render(request, 'hrd/jatah_cuti_grid.html', {
        'data': data,
        'tahun': tahun,
    })

@require_POST
@login_required
def update_detail_jatah_cuti(request):
    if request.user.role != 'HRD':
        return JsonResponse({'success': False, 'message': 'Unauthorized'}, status=403)

    detail_id = request.POST.get('detail_id')
    dipakai = request.POST.get('dipakai') == 'true'
    jumlah_hari = int(request.POST.get('jumlah_hari', 0))
    keterangan = request.POST.get('keterangan', '')

    try:
        detail = DetailJatahCuti.objects.get(id=detail_id)
        detail.dipakai = dipakai
        detail.jumlah_hari = jumlah_hari
        detail.keterangan = keterangan
        detail.save()

        return JsonResponse({
            'success': True,
            'message': 'Berhasil disimpan',
            'dipakai': detail.dipakai,
            'jumlah_hari': detail.jumlah_hari,
            'keterangan': detail.keterangan
        })

    except DetailJatahCuti.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Data tidak ditemukan'})


@require_POST
@login_required
def edit_detail_jatah_cuti(request):
    if request.user.role != 'HRD':
        return JsonResponse({'success': False})

    jatah_id = request.POST.get('jatah_id')
    bulan = int(request.POST.get('bulan'))
    dipakai = request.POST.get('dipakai') == 'true'
    jumlah_hari = int(request.POST.get('jumlah_hari', 0))
    keterangan = request.POST.get('keterangan', '')

    try:
        jatah = JatahCuti.objects.get(id=jatah_id)
    except JatahCuti.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Jatah cuti tidak ditemukan'})

    detail, created = DetailJatahCuti.objects.get_or_create(
        jatah_cuti=jatah,
        bulan=bulan,
        tahun=jatah.tahun,
        defaults={
            'dipakai': dipakai,
            'jumlah_hari': jumlah_hari,
            'keterangan': keterangan
        }
    )

    detail.dipakai = dipakai
    detail.jumlah_hari = jumlah_hari
    detail.keterangan = keterangan
    detail.save()

    return JsonResponse({'success': True})

